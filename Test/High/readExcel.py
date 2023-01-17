# 开发者：Annona
# 开发时间：2023/1/16 13:26
import openpyxl
# 使用load_workbook()方法打开文件
book=openpyxl.load_workbook('G:/LWY_Other/testdate/python/excel/sample.xlsx')
sheet=book.active
# a1=sheet.cell(row=3,column=1).value
# a2=sheet['B2'].value
# print(a1)
# print(a2)
# cells=sheet['A1':'C6']
# for c1,c2,c3 in cells:
#     # format()功能用于在控制台上整洁地输出数据。c1,c2,c3分别表示第1-3列
#     print('{0:1} {1:8} {2:8}'.format(c1.value,c2.value,c3.value))
# 该示例逐行遍历数据,我们提供了遍历的边界
# for row in sheet.iter_rows(min_row=1,min_col=1,max_row=6,max_col=3):
#     for cell in row:
#         print(cell.value,end=' ')
#     print()
# 该示例逐列遍历数据,我们提供了遍历的边界,边界值也可以不设置
for col in sheet.iter_cols(min_row=1,min_col=1,max_row=6,max_col=3):
    for cell in col:
        print(cell.value,end=' ')
    print()
