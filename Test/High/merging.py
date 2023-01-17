# 开发者：Annona
# 开发时间：2023/1/16 16:36
from openpyxl import Workbook
from openpyxl.styles import Alignment

book = Workbook()
sheet = book.active
sheet.merge_cells('A1:B2')
cell = sheet.cell(row=1 , column=1)
cell.value = 'Test Merging'
cell.alignment=Alignment(horizontal='center', vertical='center')
book.save('G:/LWY_Other/testdate/python/excel/merging.xlsx')