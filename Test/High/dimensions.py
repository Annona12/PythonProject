# 开发者：Annona
# 开发时间：2023/1/16 15:31
from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
sheet['A3'] = 39
sheet['B3'] = 19

rows = [
    (88, 46),
    (89, 38),
    (23, 59),
    (56, 21),
    (24, 18),
    (34, 15)
]
for row in rows:
    sheet.append(row)

# dimensions属性返回非空单元格区域的左上角和右下角单元格。
print(sheet.dimensions)
# 使用min_row和max_row属性，我们可以获得包含数据的最小和最大行/列。
print("Minimum row: {0}".format(sheet.min_row))
print("Maximum row: {0}".format(sheet.max_row))
print("Minimum column: {0}".format(sheet.min_column))
print("Maximum column: {0}".format(sheet.max_column))

wb.save('G:/LWY_Other/testdate/python/excel/dimensions.xlsx')
