# 开发者：Annona
# 开发时间：2023/1/16 17:04
from openpyxl import Workbook

book=Workbook()
sheet=book.active

rows = (
    (34, 26),
    (88, 36),
    (24, 29),
    (15, 22),
    (56, 13),
    (76, 18)
)
for row in rows:
    sheet.append(row)
cell1 = sheet.cell(row=7, column=1, value='SUM')
cell2 = sheet.cell(row=7, column=2)
#将计算公式写入单元格
cell2.value = "=SUM(A1:B6)"
#更改字体样式
# cell2.font = cell2.font.copy(bold=True)
book.save('G:/LWY_Other/testdate/python/excel/formulas.xlsx')
