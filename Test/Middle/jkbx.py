# 开发者：Annona
# 开发时间：2023/1/12 13:50
# import pandas as pd
import openpyxl

book=openpyxl.load_workbook('G:/LWY_Other/testdate/python/jkbx/Excel.xlsx')
sheet=book.active

print(sheet.dimensions)
print("Minimum row: {0}".format(sheet.min_row))
print("Maximum row: {0}".format(sheet.max_row))
print("Minimum column: {0}".format(sheet.min_column))
print("Maximum column: {0}".format(sheet.max_column))
# cols=sheet.columns
# for col in cols:
#     for cell in col:
#         print(cell.value)
# for col in sheet.iter_cols():
#     for cell in col:
#         print(cell.value,end=' ')
#     print()

# df=pd.read_excel(r"G:\LWY_Other\testdate\python\jkbx\Excel.xls")
# df=pd.read_excel(r"G:\LWY_Other\testdate\python\jkbx\Excel.xls",sheet_name='Excel1')

# print(df[0])
# print(df)
# print(df.head())
# path=pd.ExcelFile(r"G:\LWY_Other\testdate\python\jkbx\Excel.xls")
# df1=pd.read_excel(path,'Excel')
# df2=pd.read_excel(path,'Excel1')
# print(df2)