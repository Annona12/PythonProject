# 开发者：Annona
# 开发时间：2023/1/16 15:44
import openpyxl
book=openpyxl.load_workbook('G:/LWY_Other/testdate/python/excel/sheets.xlsx')
#使用create_sheet()方法创建一个新图纸
book.create_sheet("Sheet4")
# get_sheet_names()方法返回工作簿中可用工作表的名称
# print(book.get_sheet_names())
print(book.sheetnames)
active_sheet=book.active
# 获取活动表并将其类型打印到终端
print(type(active_sheet))
#删除一个sheet
sheet1=book['Sheet1']
# 使用remove_sheet()方法将纸张取出
# book.remove_sheet(sheet1)
book.remove(sheet1)
print(book.sheetnames)
# 可以在指定位置创建一个新图纸。 在我们的例子中，我们在索引为 0 的位置创建一个新工作表。
book.create_sheet("Sheet5", 0)
print(book.sheetnames)
# 我们使用get_sheet_by_name()方法获得对工作表的引用。
# sheet = book.get_sheet_by_name("March")
sheet=book['Sheet3']
#修改工作表的背景颜色
sheet.sheet_properties.tabColor = "0072BA"
print(sheet.title)
print(sheet['A1'].value)
book.save('G:/LWY_Other/testdate/python/excel/sheets.xlsx')

