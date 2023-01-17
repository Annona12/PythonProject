# 开发者：Annona
# 开发时间：2023/1/16 10:56
#导入Workbook类
from openpyxl import Workbook
import time
# 我们创建一个新的工作簿。始终使用至少一个工作表创建一个工作簿
book=Workbook()
# 获得对活动工作表的引用
sheet=book.active
# # 我们将数值数据写入单元格 A1 和 A2。
# sheet['A1']='姓名'
# sheet['B1']='性别'
# sheet['A2']='Annona'
# sheet['B2']='女'
# # row是行，column是列
# sheet.cell(row=3,column=1).value='Kawi'
# sheet.cell(row=3,column=2).value='男'
# 使用save()方法将内容写入sample.xlsx文件
# rows = (
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# )
# for row in rows:
#     sheet.append(row)
# book.save('G:/LWY_Other/testdate/python/excel/sample.xlsx')

